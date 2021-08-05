from NaverNewsCrawler import NaverNewsCrawler

####사용자로 부터 기사 수집을 원하는 키워드를 input을 이용해 입력받아 ? 부분에 넣으세요
keyword = input('검색하고자 하는 keyword : ')
crawler = NaverNewsCrawler(keyword)

#### 수집한 데이터를 저장할 엑셀 파일명을 input을 이용해 입력받아 ? 부분에 넣으세요
crawler.get_news('crawling_result.xlsx')

#### 아래코드를 실행해 이메일 발송 기능에 필요한 모듈을 임포트하세요.
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import re

#### gmail 발송 기능에 필요한 계정 정보를 아래 코드에 입력하세요.
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SMTP_USER = ''
SMTP_PASSWORD = ''

#### 아래 코드를 실행해 메일 발송에 필요한 send_mail 함수를 만드세요.
def send_mail(name, addr, subject, contents, attachment=None):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment; filename="' + filename + '"')
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()
    print('메일 전송을 성공적으로 마쳤습니다.')


#### 엑셀 파일의 정보를 읽어올 수 있는 모듈을 import하세요.
#### 프로젝트 폴더에 있는 email_list.xlsx 파일에 이메일 받을 사람들의 정보를 읽어오세요.
def get_receiver(list_file) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(list_file)
    data = wb.active
    length = 'C' + str(data.max_row)
    data_area = data['B3:' + length]

    result = {}
    for row in data_area:
        temp = []
        for cell in row:
            temp.append(cell.value)

        receiver_name = temp[0]
        receiver_email = temp[1]

        # 동명이인이 없다고 가정하고 코드를 작성해보았습니다.
        if receiver_name != None and receiver_email != None:
            result[receiver_name] = receiver_email

    return result


def send_all_mails(receiver_list: dict):
    for name in receiver_list.keys():
        address = receiver_list[name]
        subject = '검색하신 키워드 "' + keyword + '"에 대한 결과 안내 메일'
        contents = '안녕하세요 ' + name + '님!\n\n검색하신 키워드인 "' + keyword + '"에 대한 기사 내용을 첨부 파일에 담아 메일 드립니다.'
        attachment = 'crawling_result.xlsx'
        send_mail(name, address, subject, contents, attachment)



# 이메일 발송 자동화 프로그램
if __name__ == '__main__':
    SMTP_USER = input("Your Gmail Address : ")
    SMTP_PASSWORD = input("Password : ")

    email_list_file = 'email_list.xlsx'
    receiver_list = get_receiver(email_list_file)

    send_all_mails(receiver_list)